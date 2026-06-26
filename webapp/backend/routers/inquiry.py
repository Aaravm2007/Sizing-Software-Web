import io
import sys
from pathlib import Path
from typing import Any, Dict

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

sys.path.insert(0, str(Path(__file__).parent.parent))
from auth import get_current_user, get_expert_user
from inquiry_db import (
    push_row, list_rows, update_row, delete_row, init_inquiry_db,
    suggest_next_inquiry_code, list_inquiry_page, count_inquiry, export_inquiry_rows,
    _COLS as _INQ_COLS,
)
from user_db import get_user_inquiry_db

init_inquiry_db()
router = APIRouter()

_EXPORT_SKIP = {"created_at"}
_EXPORT_HEADERS = [c for c in _INQ_COLS if c not in _EXPORT_SKIP]

_LABEL_MAP = {
    "sr_no": "Sr No", "inquiry_code": "Inquiry Code", "inquiry_date": "Inquiry Date",
    "type": "Type", "sales_person": "Sales Person", "solution_provider": "Solution Provider",
    "project_customer": "Project / Customer", "ups_make": "UPS Make", "ups_model": "UPS Model",
    "ups_kva": "UPS (KVA)", "actual_load_kva": "Load (KVA)", "load_kw": "Load (KW)",
    "power_factor": "Power Factor", "inverter_efficiency": "Inv. Eff (%)",
    "dc_voltage": "DC Voltage", "backup_min": "Backup (min)", "cell_chemistry": "Chemistry",
    "ageing_pct": "Ageing (%)", "design_margin_pct": "Design Margin (%)",
    "dod_margin_pct": "DOD Margin (%)", "derating_pct": "Derating (%)",
    "capacity_ah": "Capacity (AH)", "centre_tap": "Centre Tap",
    "cell_type": "Cell Type", "ageing_type": "Ageing Type", "backup_time_min": "Backup Time (min)",
    "part_code": "Part Code", "qty_system": "Qty System", "rate_system": "Rate/System",
    "price_system": "Price/System", "rack_dim": "Rack Dim", "qty": "Qty",
    "per_rack_price": "Per Rack Price", "price": "Price",
    "custom_cost_desc": "Custom Cost Desc", "custom_cost_price": "Custom Cost Price",
    "rack1_dim": "Rack1 Dim", "rack1_qty": "Rack1 Qty", "rack1_rate": "Rack1 Rate", "rack1_price": "Rack1 Price",
    "rack2_dim": "Rack2 Dim", "rack2_qty": "Rack2 Qty", "rack2_rate": "Rack2 Rate", "rack2_price": "Rack2 Price",
    "cc1_desc": "CC1 Desc", "cc1_price": "CC1 Price",
    "cc2_desc": "CC2 Desc", "cc2_price": "CC2 Price",
    "cc3_desc": "CC3 Desc", "cc3_price": "CC3 Price",
    "cc4_desc": "CC4 Desc", "cc4_price": "CC4 Price",
    "cc5_desc": "CC5 Desc", "cc5_price": "CC5 Price",
    "datasheet": "Datasheet", "sizing_sheet": "Sizing Sheet", "gad": "GAD",
    "battery_compliance": "Bat. Compliance", "warranty": "Warranty (yrs)",
    "remarks": "Remarks", "handled_by": "Handled By",
    "submission_date": "Submission Date", "submitted_to": "Submitted To",
    "submitted_by": "Submitted By", "quote_code": "Quote Code", "sol_no": "Sol No",
    "dollar_rate": "Dollar Rate", "base_partcode": "Base Part Code", "quote_format": "Quote Format",
}


def _parse_filter_params(request: Request):
    params = dict(request.query_params)
    page  = max(1, int(params.pop("page", 1)))
    limit = min(500, max(1, int(params.pop("limit", 250))))
    search = params.pop("search", "")
    fields = {k: v for k, v in params.items() if v}
    return page, limit, search, fields


class InquiryEntry(BaseModel):
    inquiry_code: str = ""
    inquiry_date: str = ""
    type: str = ""
    sales_person: str = ""
    solution_provider: str = ""
    project_customer: str = ""
    ups_make: str = ""
    ups_model: str = ""
    ups_kva: str = ""
    actual_load_kva: str = ""
    load_kw: str = ""
    power_factor: str = ""
    inverter_efficiency: str = ""
    dc_voltage: str = ""
    backup_min: str = ""
    cell_chemistry: str = ""
    ageing_pct: str = ""
    design_margin_pct: str = ""
    dod_margin_pct: str = ""
    derating_pct: str = ""
    capacity_ah: str = ""
    part_code: str = ""
    qty: str = ""
    centre_tap: str = ""
    cell_type: str = ""
    qty_system: str = ""
    rate_system: str = ""
    price_system: str = ""
    submission_date: str = ""
    submitted_to: str = ""
    price: str = ""
    per_rack_price: str = ""
    rack_dim: str = ""
    custom_cost_desc: str = ""
    custom_cost_price: str = ""
    rack1_dim: str = ""
    rack1_qty: str = ""
    rack1_rate: str = ""
    rack1_price: str = ""
    rack2_dim: str = ""
    rack2_qty: str = ""
    rack2_rate: str = ""
    rack2_price: str = ""
    cc1_desc: str = ""
    cc1_price: str = ""
    cc2_desc: str = ""
    cc2_price: str = ""
    cc3_desc: str = ""
    cc3_price: str = ""
    cc4_desc: str = ""
    cc4_price: str = ""
    cc5_desc: str = ""
    cc5_price: str = ""
    datasheet: str = "NO"
    sizing_sheet: str = "YES"
    gad: str = "NO"
    battery_compliance: str = "NO"
    warranty: str = "5 year"
    remarks: str = ""
    handled_by: str = ""


@router.get("/next-inquiry-code")
def next_inquiry_code(_=Depends(get_current_user)):
    return suggest_next_inquiry_code()


@router.get("/mine")
def list_my_entries(user=Depends(get_current_user)):
    db = get_user_inquiry_db(user["username"])
    init_inquiry_db(db)
    return list_rows(db_path=db)


@router.get("/export-excel")
def export_excel(request: Request, _=Depends(get_expert_user)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    _, _, search, fields = _parse_filter_params(request)
    rows = export_inquiry_rows(search, fields)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiry Sheet"

    labels = [_LABEL_MAP.get(h, h) for h in _EXPORT_HEADERS]
    ws.append(labels)

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 28

    for row in rows:
        ws.append([str(row.get(h, "") or "") for h in _EXPORT_HEADERS])

    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 8), 30)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inquiry_sheet.xlsx"},
    )


@router.get("")
def list_entries(request: Request, _=Depends(get_current_user)):
    page, limit, search, fields = _parse_filter_params(request)
    total = count_inquiry(search, fields)
    rows  = list_inquiry_page(page, limit, search, fields)
    pages = max(1, (total + limit - 1) // limit)
    return {"rows": rows, "total": total, "pages": pages}


@router.post("", status_code=201)
def create_entry(body: InquiryEntry, user=Depends(get_current_user)):
    data = body.dict()
    data["handled_by"] = user["username"]
    sr = push_row(data)
    return {"sr_no": sr}


@router.patch("/{sr_no}")
def update_entry(sr_no: int, body: Dict[str, Any], _=Depends(get_expert_user)):
    update_row(sr_no, body)
    return {"ok": True}


@router.delete("/{sr_no}")
def delete_entry(sr_no: int, _=Depends(get_expert_user)):
    delete_row(sr_no)
    return {"ok": True}
