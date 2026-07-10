import sys
import os
import math
import tempfile
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import FileResponse
from pydantic import BaseModel

APP_DIR = Path(__file__).parent.parent.parent.parent
BACKEND_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(APP_DIR))
sys.path.insert(0, str(BACKEND_DIR))

from sql_handler import (
    init_sizing_db, fetch_all_projects, fetch_all_sizings, fetch_sizing_by_sr,
    fetch_max_sr_no, insert_sizing, update_sizing, delete_sizing, delete_project,
    duplicate_sizing,
)
from auth import get_current_user
from user_db import get_user_sizing_db, get_user_temp_db

router = APIRouter()

# ── Pydantic models ───────────────────────────────────────────────────────────

class ProjectCreate(BaseModel):
    name: str

class SizingData(BaseModel):
    customer_name: str = ""
    solution_provider: str = ""
    ups_make: str = ""
    ups_model: str = ""
    ups_rating_kva: float = 0
    actual_load_kva: float = 0
    actual_load_kw: float = 0
    power_factor: float = 0
    inverter_efficiency: float = 0
    nominal_dc_voltage: float = 0
    backup_requirement_min: float = 0
    ageing_percent: float = 0
    design_margin_percent: float = 0
    dod_margin_percent: float = 0
    derating_factor_percent: float = 0
    number_of_cells: int = 0
    cell_chemistry: str = "LFP"
    calculated_load_kw: float = 0
    max_charging_voltage: float = 0
    end_cell_voltage: float = 0
    energy_required_kwh: float = 0
    capacity_required_ah: float = 0
    cap_with_ageing_ah: float = 0
    cap_with_design_margin_ah: float = 0
    cap_with_dod_margin_ah: float = 0
    cap_with_derating_factor_ah: float = 0
    nearest_capacity_ah: float = 0
    offered_battery_config: str = ""
    total_available_energy_kwh: float = 0
    backup_time_min: float = 0
    ageing_type: str = "BOL"

# ── helpers ───────────────────────────────────────────────────────────────────

def _to_db_dict(d: SizingData) -> dict:
    return {
        "Customer Name": d.customer_name,
        "Solution Provider": d.solution_provider,
        "UPS Make": d.ups_make,
        "UPS Model": d.ups_model,
        "UPS Rating (KVA)": d.ups_rating_kva,
        "Actual Load (KVA)": d.actual_load_kva,
        "Actual Load (kW)": d.actual_load_kw,
        "Power Factor": d.power_factor,
        "Inverter Efficiency": d.inverter_efficiency,
        "Nominal DC Voltage (V)": d.nominal_dc_voltage,
        "Backup Requirement (Min)": d.backup_requirement_min,
        "Ageing (%)": d.ageing_percent,
        "Design Margin (%)": d.design_margin_percent,
        "DOD Margin (%)": d.dod_margin_percent,
        "Derating Factor (%)": d.derating_factor_percent,
        "Number of Cells": d.number_of_cells,
        "Cell Chemistry": d.cell_chemistry,
        "Calculated Load (kW)": d.calculated_load_kw,
        "Max Charging Voltage (V)": d.max_charging_voltage,
        "End Cell Voltage (V)": d.end_cell_voltage,
        "Energy Required (kWh)": d.energy_required_kwh,
        "Capacity Required (Ah)": d.capacity_required_ah,
        "Cap req w/ Ageing (Ah)": d.cap_with_ageing_ah,
        "Cap req w/ Design Margin (Ah)": d.cap_with_design_margin_ah,
        "Cap req w/ DOD (Ah)": d.cap_with_dod_margin_ah,
        "Cap req w/ Derating (Ah)": d.cap_with_derating_factor_ah,
        "Nearest Available Capacity (Ah)": d.nearest_capacity_ah,
        "Offered Battery Configuration": d.offered_battery_config,
        "Total Available Energy (kWh)": d.total_available_energy_kwh,
        "Backup Time (Min)": d.backup_time_min,
        "Ageing Type": d.ageing_type,
    }

def _row_to_dict(row) -> dict:
    if row is None:
        return {}
    return {
        "sr_no": row[0],
        "customer_name": row[1] or "",
        "solution_provider": row[2] or "",
        "ups_make": row[3] or "",
        "ups_model": row[4] or "",
        "ups_rating_kva": row[5] or 0,
        "actual_load_kva": row[6] or 0,
        "actual_load_kw": row[7] or 0,
        "power_factor": row[8] or 0,
        "inverter_efficiency": row[9] or 0,
        "nominal_dc_voltage": row[10] or 0,
        "backup_requirement_min": row[11] or 0,
        "ageing_percent": row[12] or 0,
        "design_margin_percent": row[13] or 0,
        "dod_margin_percent": row[14] or 0,
        "derating_factor_percent": row[15] or 0,
        "number_of_cells": row[16] or 0,
        "cell_chemistry": row[17] or "LFP",
        "calculated_load_kw": row[18] or 0,
        "max_charging_voltage": row[19] or 0,
        "end_cell_voltage": row[20] or 0,
        "energy_required_kwh": row[21] or 0,
        "capacity_required_ah": row[22] or 0,
        "cap_with_ageing_ah": row[23] or 0,
        "cap_with_design_margin_ah": row[24] or 0,
        "cap_with_dod_margin_ah": row[25] or 0,
        "cap_with_derating_factor_ah": row[26] or 0,
        "nearest_capacity_ah": row[27] or 0,
        "offered_battery_config": row[28] or "",
        "total_available_energy_kwh": row[29] or 0,
        "backup_time_min": row[30] or 0,
        "ageing_type": row[31] if len(row) > 31 else "BOL",
    }


# ── project endpoints ─────────────────────────────────────────────────────────

@router.get("/projects")
def list_projects(user=Depends(get_current_user)):
    db = get_user_sizing_db(user["username"])
    projects = fetch_all_projects(db_path=db)
    return [{"name": p, "count": fetch_max_sr_no(p, db_path=db)} for p in projects]


@router.post("/projects", status_code=201)
def create_project(body: ProjectCreate, user=Depends(get_current_user)):
    if not body.name.strip():
        raise HTTPException(400, "Project name required")
    try:
        init_sizing_db(body.name.strip(), db_path=get_user_sizing_db(user["username"]))
    except Exception as e:
        raise HTTPException(500, str(e))
    return {"name": body.name.strip()}


@router.delete("/projects/{name}")
def remove_project(name: str, user=Depends(get_current_user)):
    try:
        delete_project(name, db_path=get_user_sizing_db(user["username"]))
    except Exception as e:
        raise HTTPException(500, str(e))
    return {"detail": "deleted"}


# ── sizing CRUD ───────────────────────────────────────────────────────────────

@router.get("/projects/{name}/sizings")
def list_sizings(name: str, user=Depends(get_current_user)):
    db = get_user_sizing_db(user["username"])
    try:
        rows = fetch_all_sizings(name, db_path=db)
    except ValueError:
        raise HTTPException(404, f"Project '{name}' not found")
    return [{"sr_no": r[0], "offered_battery_config": r[1]} for r in rows]


@router.get("/projects/{name}/sizings/{sr_no}")
def get_sizing(name: str, sr_no: int, user=Depends(get_current_user)):
    db = get_user_sizing_db(user["username"])
    try:
        row = fetch_sizing_by_sr(name, sr_no, db_path=db)
    except ValueError:
        raise HTTPException(404, f"Project '{name}' not found")
    if row is None:
        raise HTTPException(404, "Sizing not found")
    return _row_to_dict(row)


@router.post("/projects/{name}/sizings", status_code=201)
def add_sizing(name: str, body: SizingData, user=Depends(get_current_user)):
    db = get_user_sizing_db(user["username"])
    try:
        sr_no = insert_sizing(name, _to_db_dict(body), db_path=db)
    except ValueError as e:
        raise HTTPException(404, str(e))
    return {"detail": "created", "sr_no": sr_no}


@router.put("/projects/{name}/sizings/{sr_no}")
def edit_sizing(name: str, sr_no: int, body: SizingData, user=Depends(get_current_user)):
    db = get_user_sizing_db(user["username"])
    try:
        update_sizing(name, sr_no, _to_db_dict(body), db_path=db)
    except ValueError as e:
        raise HTTPException(404, str(e))
    return {"detail": "updated"}


@router.delete("/projects/{name}/sizings/{sr_no}")
def remove_sizing(name: str, sr_no: int, user=Depends(get_current_user)):
    db = get_user_sizing_db(user["username"])
    try:
        delete_sizing(name, sr_no, db_path=db)
    except ValueError as e:
        raise HTTPException(404, str(e))
    return {"detail": "deleted"}


@router.post("/projects/{name}/sizings/{sr_no}/duplicate")
def dup_sizing(name: str, sr_no: int, user=Depends(get_current_user)):
    db = get_user_sizing_db(user["username"])
    try:
        duplicate_sizing(name, sr_no, db_path=db)
    except ValueError as e:
        raise HTTPException(404, str(e))
    return {"detail": "duplicated"}


# ── export ────────────────────────────────────────────────────────────────────

SIZING_CELL_MAP = {
    "Customer Name": "C4", "Solution Provider": "C5",
    "Date": "H3",
    "UPS Make": "A8", "UPS Model": "B8",
    "UPS Rating (KVA)": "C8",
    "Power Factor": "E8", "Inverter Efficiency": "F8",
    "Nominal DC Voltage (V)": "G8", "Backup Requirement (Min)": "H8",
    "Cell Chemistry": "E11", "Calculated Load (kW)": "E12",
    "Max Charging Voltage (V)": "E13", "End Cell Voltage (V)": "E14",
    "Energy Required (kWh)": "E17", "Capacity Required_Base (Ah)": "E18",
    "Ageing (%)": "E19", "Design Margin (%)": "E20",
    "DOD Margin (%)": "E21", "Derating Factor (%)": "E22",
    "Capacity Required (Ah)": "E23", "Cap req w/ Design Margin (Ah)": "E24",
    "Cap req w/ DOD (Ah)": "E25", "Cap req w/ Derating (Ah)": "E26",
    "Nearest Available Capacity (Ah)": "E27",
    "Offered Battery Configuration": "E28",
    "Total Available Energy (kWh)": "E29", "Backup Time (Min)": "E30",
}


def _sizing_row_to_data(row) -> dict:
    from datetime import datetime as _dt
    nv = lambda v: v if v else ""
    end_v      = row[20] or 0
    energy_kwh = row[21] or 0
    base_cap   = row[22] or ""
    ageing     = row[12] or 0
    design     = row[13] or 0
    dod        = row[14] or 0
    derating   = row[15] or 0
    kva = row[6] or 0
    kw  = row[7] or 0
    if kw:
        load_label, load_value = "Actual Load (kW)", kw
    elif kva:
        load_label, load_value = "Actual Load (KVA)", kva
    else:
        load_label, load_value = "Actual Load", ""
    ageing_type = (row[31] if len(row) > 31 else None) or "BOL"
    backup_label = f"Backup Time (Min) at {ageing_type}"
    return {
        "_load_label": load_label, "_load_value": load_value,
        "_backup_label": backup_label,
        "Date": "Date: " + _dt.now().strftime("%d/%m/%Y"),
        "Customer Name": row[1], "Solution Provider": row[2],
        "UPS Make": row[3], "UPS Model": row[4],
        "UPS Rating (KVA)": nv(row[5]),
        "Power Factor": nv(row[8]), "Inverter Efficiency": nv(row[9]),
        "Nominal DC Voltage (V)": nv(row[10]), "Backup Requirement (Min)": nv(row[11]),
        "Ageing (%)": ageing / 100.0 if ageing else "",
        "Design Margin (%)": design / 100.0 if design else "",
        "DOD Margin (%)": dod / 100.0 if dod else "",
        "Derating Factor (%)": derating / 100.0 if derating else "",
        "Cell Chemistry": row[17], "Calculated Load (kW)": nv(row[18]),
        "Max Charging Voltage (V)": nv(row[19]), "End Cell Voltage (V)": nv(end_v),
        "Energy Required (kWh)": nv(energy_kwh),
        "Capacity Required_Base (Ah)": base_cap,
        "Capacity Required (Ah)": nv(row[23]),
        "Cap req w/ Design Margin (Ah)": nv(row[24]),
        "Cap req w/ DOD (Ah)": nv(row[25]),
        "Cap req w/ Derating (Ah)": nv(row[26]),
        "Nearest Available Capacity (Ah)": nv(row[27]),
        "Offered Battery Configuration": row[28] or "",
        "Total Available Energy (kWh)": nv(row[29]),
        "Backup Time (Min)": nv(row[30]),
    }


def _copy_images(src_ws, dst_ws):
    """Copy embedded images from src_ws to dst_ws (copy_worksheet drops them)."""
    from io import BytesIO
    from copy import deepcopy
    from openpyxl.drawing.image import Image as XLImage
    for img in getattr(src_ws, "_images", []):
        try:
            ref = img.ref
            if hasattr(ref, "read"):
                ref.seek(0)
                data = ref.read()
                ref.seek(0)
            else:
                with open(ref, "rb") as f:
                    data = f.read()
            new_img = XLImage(BytesIO(data))
            new_img.anchor = deepcopy(img.anchor)
            dst_ws.add_image(new_img)
        except Exception:
            pass


def _dash(v):
    """Return '-' for any blank/zero value so exported cells never show empty."""
    if v is None or v == "" or v == 0 or v == 0.0:
        return "-"
    return v


def _write_cell(ws, addr: str, value):
    """Write value and force black font so theme-based light colors don't hide text."""
    from openpyxl.styles import Font
    from copy import copy as _copy
    cell = ws[addr]
    cell.value = value
    existing = cell.font
    cell.font = Font(
        name=existing.name, size=existing.size, bold=existing.bold,
        italic=existing.italic, underline=existing.underline,
        strike=existing.strike, color="FF000000",
    )


def _build_excel(name: str, sr_no: Optional[int], db_path: str = None) -> str:
    import openpyxl

    template_path = str(APP_DIR / "templates" / "Sizing_template.xlsx")
    records = [(sr_no,)] if sr_no else [(r[0],) for r in fetch_all_sizings(name, db_path=db_path)]

    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb.active
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    for sn, in records:
        row  = fetch_sizing_by_sr(name, sn, db_path=db_path)
        data = _sizing_row_to_data(row)
        ws = out_wb.create_sheet(title=f"Sizing {sn}")
        # copy dimensions, styles, merges from template
        for col, dim in template_ws.column_dimensions.items():
            ws.column_dimensions[col].width = dim.width
        for row_n, dim in template_ws.row_dimensions.items():
            ws.row_dimensions[row_n].height = dim.height
        from copy import copy as _copy
        for r in template_ws.iter_rows():
            for cell in r:
                nc = ws[cell.coordinate]
                nc.value = cell.value
                if cell.has_style:
                    nc.font = _copy(cell.font)
                    nc.border = _copy(cell.border)
                    nc.fill = _copy(cell.fill)
                    nc.number_format = _copy(cell.number_format)
                    nc.protection = _copy(cell.protection)
                    nc.alignment = _copy(cell.alignment)
        for mr in template_ws.merged_cells.ranges:
            ws.merge_cells(str(mr))
        _copy_images(template_ws, ws)
        for key, cell_addr in SIZING_CELL_MAP.items():
            _write_cell(ws, cell_addr, _dash(data.get(key, "")))
        _write_cell(ws, "D7", data["_load_label"])
        _write_cell(ws, "D8", _dash(data["_load_value"]))
        _write_cell(ws, "D30", data["_backup_label"])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    out_wb.save(tmp.name)
    return tmp.name


@router.get("/projects/{name}/export/excel")
def export_excel(name: str, sr_no: Optional[int] = None, user=Depends(get_current_user)):
    sdb = get_user_sizing_db(user["username"])
    try:
        path = _build_excel(name, sr_no, db_path=sdb)
    except Exception as e:
        raise HTTPException(500, str(e))

    # DISABLED: inquiry auto-sync replaced by pending export history
    # if sr_no:
    #     try:
    #         import time as _time
    #         from inquiry_db import push_row as _push_inq
    #         srow = fetch_sizing_by_sr(name, sr_no, db_path=sdb)
    #         if srow:
    #             _push_inq({
    #                 "inquiry_date": _time.strftime("%d/%m/%Y"),
    #                 "type": "Sizing", "sales_person": "",
    #                 "solution_provider": str(srow[2] or ""),
    #                 "project_customer": str(srow[1] or ""),
    #                 "ups_make": str(srow[3] or ""), "ups_model": str(srow[4] or ""),
    #                 "ups_kva": str(srow[5] or ""),
    #                 "actual_load_kva": str(srow[6] or ""), "load_kw": str(srow[7] or ""),
    #                 "power_factor": str(srow[8] or ""), "inverter_efficiency": str(srow[9] or ""),
    #                 "dc_voltage": str(srow[10] or ""), "backup_min": str(srow[11] or ""),
    #                 "cell_chemistry": str(srow[17] or ""),
    #                 "ageing_pct": str(srow[12] or ""), "design_margin_pct": str(srow[13] or ""),
    #                 "dod_margin_pct": str(srow[14] or ""), "derating_pct": str(srow[15] or ""),
    #                 "capacity_ah": str(srow[27] or ""),
    #                 "datasheet": "NO", "sizing_sheet": "YES", "gad": "NO",
    #             })
    #     except Exception:
    #         pass

    fname = f"{name}_sizing{'_'+str(sr_no) if sr_no else '_all'}.xlsx"
    return FileResponse(path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename=fname, background=None)


class WizardCol(BaseModel):
    ups_make: str = ""
    ups_model: str = ""
    ups_rating_kva: str = ""
    actual_load_kva: str = ""
    actual_load_kw: str = ""
    power_factor: str = ""
    inverter_efficiency: str = ""
    nominal_dc_voltage: str = ""
    backup_requirement_min: str = ""
    ageing_type: str = "BOL"
    ageing_percent: str = ""
    design_margin_percent: str = ""
    dod_margin_percent: str = ""
    derating_factor_percent: str = ""
    cell_type: str = ""
    calculated_load_kw: str = ""
    max_charging_voltage: str = ""
    end_cell_voltage: str = ""
    energy_required_kwh: str = ""
    total_available_energy_kwh: str = ""
    capacity_required_ah: str = ""
    cap_with_ageing_ah: str = ""
    cap_with_design_margin_ah: str = ""
    cap_with_dod_margin_ah: str = ""
    cap_with_derating_ah: str = ""
    nearest_capacity_ah: str = ""
    offered_battery_config: str = ""
    backup_time_min: str = ""

class WizardExportBody(BaseModel):
    project_name: str = ""
    customer_name: str = ""
    solution_provider: str = ""
    cols: list[WizardCol]


def _build_wizard_excel(body: "WizardExportBody") -> str:
    import openpyxl

    def num(v: str):
        if not v: return ""
        try: return float(v)
        except: return v

    def pct(v: str) -> float:
        try: return float(v) / 100.0 if v else 0.0
        except: return 0.0

    template_path = str(APP_DIR / "templates" / "Sizing_template.xlsx")
    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb.active
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    from datetime import datetime as _dt
    from copy import copy as _copy
    for idx, col in enumerate(body.cols, 1):
        w_kw  = num(col.actual_load_kw)
        w_kva = num(col.actual_load_kva)
        if w_kw:
            w_load_label, w_load_value = "Actual Load (kW)", w_kw
        elif w_kva:
            w_load_label, w_load_value = "Actual Load (KVA)", w_kva
        else:
            w_load_label, w_load_value = "Actual Load", ""
        data = {
            "Date": "Date: " + _dt.now().strftime("%d/%m/%Y"),
            "Customer Name": body.customer_name,
            "Solution Provider": body.solution_provider,
            "UPS Make": col.ups_make,
            "UPS Model": col.ups_model,
            "UPS Rating (KVA)": num(col.ups_rating_kva) or "",
            "Power Factor": num(col.power_factor) or "",
            "Inverter Efficiency": num(col.inverter_efficiency) or "",
            "Nominal DC Voltage (V)": num(col.nominal_dc_voltage) or "",
            "Backup Requirement (Min)": num(col.backup_requirement_min) or "",
            "Cell Chemistry": col.cell_type,
            "Calculated Load (kW)": num(col.calculated_load_kw) or "",
            "Max Charging Voltage (V)": num(col.max_charging_voltage) or "",
            "End Cell Voltage (V)": num(col.end_cell_voltage) or "",
            "Energy Required (kWh)": num(col.energy_required_kwh) or "",
            "Total Available Energy (kWh)": num(col.total_available_energy_kwh) or "",
            "Capacity Required_Base (Ah)": num(col.capacity_required_ah) or "",
            "Ageing (%)": pct(col.ageing_percent) or "",
            "Design Margin (%)": pct(col.design_margin_percent) or "",
            "DOD Margin (%)": pct(col.dod_margin_percent) or "",
            "Derating Factor (%)": pct(col.derating_factor_percent) or "",
            "Capacity Required (Ah)": num(col.cap_with_ageing_ah) or "",
            "Cap req w/ Design Margin (Ah)": num(col.cap_with_design_margin_ah) or "",
            "Cap req w/ DOD (Ah)": num(col.cap_with_dod_margin_ah) or "",
            "Cap req w/ Derating (Ah)": num(col.cap_with_derating_ah) or "",
            "Nearest Available Capacity (Ah)": num(col.nearest_capacity_ah) or "",
            "Offered Battery Configuration": col.offered_battery_config,
            "Backup Time (Min)": num(col.backup_time_min) or "",
        }
        ws = out_wb.create_sheet(title=f"Sizing {idx}")
        for c, dim in template_ws.column_dimensions.items():
            ws.column_dimensions[c].width = dim.width
        for rn, dim in template_ws.row_dimensions.items():
            ws.row_dimensions[rn].height = dim.height
        for r in template_ws.iter_rows():
            for cell in r:
                nc = ws[cell.coordinate]
                nc.value = cell.value
                if cell.has_style:
                    nc.font = _copy(cell.font)
                    nc.border = _copy(cell.border)
                    nc.fill = _copy(cell.fill)
                    nc.number_format = _copy(cell.number_format)
                    nc.protection = _copy(cell.protection)
                    nc.alignment = _copy(cell.alignment)
        for mr in template_ws.merged_cells.ranges:
            ws.merge_cells(str(mr))
        _copy_images(template_ws, ws)
        for key, cell_addr in SIZING_CELL_MAP.items():
            _write_cell(ws, cell_addr, _dash(data.get(key, "")))
        _write_cell(ws, "D7", w_load_label)
        _write_cell(ws, "D8", _dash(w_load_value))
        _write_cell(ws, "D30", f"Backup Time (Min) at {col.ageing_type or 'BOL'}")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    out_wb.save(tmp.name)
    return tmp.name


# ── Export as Project (Multi Format) ───────────────────────────────────────────
# Isolated from WizardCol/WizardExportBody/_build_wizard_excel above — zero risk
# to the existing /export-wizard endpoint.

class ProjectCol(BaseModel):
    nominal_dc_voltage: str = ""
    backup_requirement_min: str = ""
    cell_chemistry: str = "LFP"
    centre_tap: str = ""
    cell_type: str = ""
    ups_rating_kva: str = ""
    power_factor: str = ""
    inverter_efficiency: str = ""
    end_cell_voltage: str = ""
    max_charging_voltage: str = ""
    calculated_load_kw: str = ""
    energy_required_kwh: str = ""
    capacity_required_ah: str = ""
    nearest_capacity_ah: str = ""
    total_available_energy_kwh: str = ""
    backup_time_min: str = ""
    offered_battery_config: str = ""
    cost: str = ""                 # row 27 — Price with Cabinet & BMS
    modular_rack: str = "-"        # row 28 label (rack key, or "-")
    modular_rack_price: str = ""   # row 28 value


class ProjectGroup(BaseModel):
    name: str
    cols: list[ProjectCol]


class ProjectExportBody(BaseModel):
    customer_name: str = ""
    solution_provider: str = ""
    quote_code: str = ""
    sales_person: str = ""
    dollar_rate: str = ""
    warranty_years: str = "5"
    groups: list[ProjectGroup]
    # raw frontend wizard state (full ColState[] + groups), stored blindly for
    # "Load Past Wizard" restore — backend never inspects its shape.
    full_state: Optional[dict] = None


# field -> template row number (Multi Format.xlsx layout, col A = label, cols B.. = solutions)
PROJECT_ROW_MAP = {
    "nominal_dc_voltage":         7,
    "backup_requirement_min":     8,
    "cell_chemistry":             9,
    "centre_tap":                10,
    "cell_type":                 11,
    "ups_rating_kva":            12,
    "power_factor":              13,
    "inverter_efficiency":       14,
    "end_cell_voltage":          15,
    "max_charging_voltage":      16,
    "calculated_load_kw":        18,
    "energy_required_kwh":       19,
    "capacity_required_ah":      20,
    "nearest_capacity_ah":       21,
    "total_available_energy_kwh": 22,
    "backup_time_min":           25,
    "offered_battery_config":    26,
    "cost":                      27,
    "modular_rack_price":        28,
}
# rows that carry per-solution data and must be cleared of template demo values
# before writing (23/24 = secondary "kwh approx" sub-calc, intentionally left blank)
PROJECT_DATA_ROWS = list(PROJECT_ROW_MAP.values()) + [23, 24]

_CENTRE_TAP_MAP = {"Centre Tap": "3Wire", "Non Centre Tap": "2Wire"}


def _sanitize_sheet_name(name: str, used: set) -> str:
    invalid = set('\\/?*[]:')
    cleaned = "".join(c for c in name if c not in invalid).strip() or "Group"
    cleaned = cleaned[:31]
    candidate = cleaned
    n = 2
    while candidate in used:
        suffix = f" ({n})"
        candidate = cleaned[:31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def _build_project_excel(body: "ProjectExportBody") -> str:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt

    # Load the template into memory once. We clone its reference sheet natively
    # (openpyxl's own copy_worksheet, not a manual attribute-by-attribute rebuild)
    # so every bit of native formatting — styles, merges, conditional formatting,
    # named styles, everything — survives untouched. The loaded workbook is only
    # ever saved to a new temp path at the end; the template file on disk is
    # never written back to.
    template_path = str(APP_DIR / "templates" / "Multi_Format_template.xlsx")
    wb = openpyxl.load_workbook(template_path)
    reference_ws = wb["Base"]  # style reference: 8 cols = exactly MAX_GROUP_SIZE (7) solution slots

    used_names: set = set()
    new_sheets = []
    for group in body.groups:
        ws = wb.copy_worksheet(reference_ws)
        ws.title = _sanitize_sheet_name(group.name, used_names)
        _copy_images(reference_ws, ws)  # copy_worksheet doesn't carry images/charts
        new_sheets.append(ws)

        # clear demo data from the template across all solution columns (B..H)
        for row_n in PROJECT_DATA_ROWS:
            for col_n in range(2, 9):
                ws.cell(row=row_n, column=col_n).value = None

        # shared header (single merged cell per row, write to the anchor cell)
        _write_cell(ws, "B4", body.customer_name)
        _write_cell(ws, "B5", body.solution_provider)
        _write_cell(ws, "H3", "Date:" + _dt.now().strftime("%d.%m.%Y"))

        for i, col in enumerate(group.cols[:7]):
            col_letter = get_column_letter(2 + i)  # B, C, D, E, F, G, H
            values = {
                "nominal_dc_voltage":         col.nominal_dc_voltage,
                "backup_requirement_min":     col.backup_requirement_min,
                "cell_chemistry":             col.cell_chemistry,
                "centre_tap":                 _CENTRE_TAP_MAP.get(col.centre_tap, col.centre_tap),
                "cell_type":                  col.cell_type,
                "ups_rating_kva":             col.ups_rating_kva,
                "power_factor":               col.power_factor,
                "inverter_efficiency":        col.inverter_efficiency,
                "end_cell_voltage":           col.end_cell_voltage,
                "max_charging_voltage":       col.max_charging_voltage,
                "calculated_load_kw":         col.calculated_load_kw,
                "energy_required_kwh":        col.energy_required_kwh,
                "capacity_required_ah":       col.capacity_required_ah,
                "nearest_capacity_ah":        col.nearest_capacity_ah,
                "total_available_energy_kwh": col.total_available_energy_kwh,
                "backup_time_min":            col.backup_time_min,
                "offered_battery_config":     col.offered_battery_config,
                "cost":                       col.cost,
                "modular_rack_price":         col.modular_rack_price if col.modular_rack != "-" else "",
            }
            for key, row_n in PROJECT_ROW_MAP.items():
                addr = f"{col_letter}{row_n}"
                _write_cell(ws, addr, _dash(values.get(key, "")))

    # drop every original template sheet — only our per-group clones ship in the output
    for name in list(wb.sheetnames):
        if wb[name] not in new_sheets:
            del wb[name]

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    return tmp.name


def _xlsx_to_pdf(xlsx_path: str) -> str:
    """Convert xlsx temp file to PDF via win32com. Deletes xlsx. Returns pdf path."""
    import win32com.client
    pdf_path = xlsx_path.replace(".xlsx", ".pdf")
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
        for sheet in wb.Sheets:
            sheet.PageSetup.PrintArea = "A1:L52"
            sheet.PageSetup.Zoom = False
            sheet.PageSetup.FitToPagesWide = 1
            sheet.PageSetup.FitToPagesTall = 1
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
    finally:
        try: wb.Close(False)
        except: pass
        excel.Quit()
    try: os.unlink(xlsx_path)
    except: pass
    return pdf_path


@router.post("/export-wizard")
def export_wizard(body: WizardExportBody, _=Depends(get_current_user)):
    try:
        xlsx_path = _build_wizard_excel(body)
    except Exception as e:
        raise HTTPException(500, str(e))
    fname = f"{body.project_name or 'wizard'}_sizing.xlsx"
    return FileResponse(xlsx_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename=fname, background=None)


@router.post("/export-wizard/pdf")
def export_wizard_pdf(body: WizardExportBody, _=Depends(get_current_user)):
    try:
        xlsx_path = _build_wizard_excel(body)
        pdf_path  = _xlsx_to_pdf(xlsx_path)
    except ImportError:
        raise HTTPException(501, "PDF export requires Microsoft Excel installed on the server")
    except Exception as e:
        raise HTTPException(500, str(e))
    fname = f"{body.project_name or 'wizard'}_sizing.pdf"
    return FileResponse(pdf_path, media_type="application/pdf", filename=fname, background=None)


@router.post("/export-project")
def export_project(body: ProjectExportBody, user=Depends(get_current_user)):
    if not body.groups or not any(g.cols for g in body.groups):
        raise HTTPException(400, "No sizings to export")
    try:
        xlsx_path = _build_project_excel(body)
    except Exception as e:
        raise HTTPException(500, str(e))

    # optional: register quote_code as metadata-only (no solutions) + Firebase backup for reload
    if body.quote_code.strip():
        from datetime import datetime as _dt
        try:
            from tempquotebase import add_new_quote
            tdb = get_user_temp_db(user["username"])
            add_new_quote(
                body.quote_code.strip(), _dt.now().strftime("%Y-%m-%d"),
                body.customer_name, body.solution_provider, "",
                db_path=tdb, sales_person=body.sales_person,
                dollar_rate=body.dollar_rate, warranty_years=body.warranty_years,
            )
        except Exception:
            pass  # quote registration is best-effort; export must still succeed

        try:
            from firebase_init import get_db
            fdb = get_db()
            fdb.reference(f"project_quotes/{body.quote_code.strip()}").set({
                "type": "Project",
                "quote_code": body.quote_code.strip(),
                "customer_name": body.customer_name,
                "solution_provider": body.solution_provider,
                "sales_person": body.sales_person,
                "dollar_rate": body.dollar_rate,
                "warranty_years": body.warranty_years,
                "created_by": user.get("username", ""),
                "created_at": int(_dt.now().timestamp() * 1000),
                "groups": [g.dict() for g in body.groups],
                "full_state": body.full_state,
            })
        except Exception:
            pass  # Firebase backup is best-effort; export must still succeed

    fname = f"{body.quote_code or 'project'}_multi_sizing.xlsx"
    return FileResponse(xlsx_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename=fname, background=None)


@router.get("/project-quotes")
def list_project_quotes(user=Depends(get_current_user)):
    """Summaries of this user's past 'Export as Project' sessions, for 'Load Past Wizard'."""
    try:
        from firebase_init import get_db
        fdb = get_db()
        all_entries = fdb.reference("project_quotes").get() or {}
    except Exception:
        return []
    username = user.get("username", "")
    out = []
    for code, entry in all_entries.items():
        if not isinstance(entry, dict) or entry.get("created_by") != username:
            continue
        out.append({
            "quote_code": entry.get("quote_code", code),
            "customer_name": entry.get("customer_name", ""),
            "solution_provider": entry.get("solution_provider", ""),
            "created_at": entry.get("created_at", 0),
        })
    out.sort(key=lambda e: e["created_at"], reverse=True)
    return out


@router.get("/project-quotes/{code}")
def get_project_quote(code: str, user=Depends(get_current_user)):
    try:
        from firebase_init import get_db
        fdb = get_db()
        entry = fdb.reference(f"project_quotes/{code}").get()
    except Exception as e:
        raise HTTPException(503, f"Firebase error: {e}")
    if not entry or not isinstance(entry, dict):
        raise HTTPException(404, "Project quote not found")
    if entry.get("created_by") != user.get("username", ""):
        raise HTTPException(403, "Not your project export")
    return entry


@router.get("/projects/{name}/export/pdf")
def export_pdf(name: str, sr_no: Optional[int] = None, _=Depends(get_current_user)):
    try:
        xlsx_path = _build_excel(name, sr_no)
        pdf_path  = _xlsx_to_pdf(xlsx_path)
    except ImportError:
        raise HTTPException(501, "PDF export requires Microsoft Excel installed on the server")
    except Exception as e:
        raise HTTPException(500, str(e))
    fname = f"{name}_sizing{'_'+str(sr_no) if sr_no else '_all'}.pdf"
    return FileResponse(pdf_path, media_type="application/pdf", filename=fname)


# ── record restore ────────────────────────────────────────────────────────────

class RestoreSizingReq(BaseModel):
    project_name: str = ""
    data: Optional[SizingData] = None       # single sizing (selected export)
    forms: Optional[list[SizingData]] = None  # multiple sizings (all export)


@router.post("/restore")
def restore_sizing(body: RestoreSizingReq, user=Depends(get_current_user)):
    if not body.project_name:
        raise HTTPException(400, "project_name is required")
    db = get_user_sizing_db(user["username"])
    try:
        existing = set(fetch_all_projects(db_path=db))
        project_name = body.project_name
        if project_name in existing:
            counter = 1
            while f"{body.project_name} ({counter})" in existing:
                counter += 1
            project_name = f"{body.project_name} ({counter})"

        init_sizing_db(project_name, db_path=db)
        if body.forms:
            for form in body.forms:
                insert_sizing(project_name, _to_db_dict(form), db_path=db)
        elif body.data:
            insert_sizing(project_name, _to_db_dict(body.data), db_path=db)
        else:
            raise HTTPException(400, "No sizing data provided")
        sr_no = fetch_max_sr_no(project_name)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    return {"project": project_name, "sr_no": sr_no}
